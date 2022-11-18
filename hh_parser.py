from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook

class HH_Parser:

    def __init__(self, vacancy_name: str, number_of_vacancy: int) -> None:
        self._vacancy_name = vacancy_name
        self._number_of_vacancy = number_of_vacancy
        self._link = 'https://hh.ru/'
        self._browser = webdriver.Chrome()
        self._browser.implicitly_wait(5)
        self._vacancy_information_list = []

    # Fill HH hello-page with vacancy name
    def _fill_start_page(self) -> None:
        self._browser.get(self._link)
        self._browser.find_element(
            By.ID, 'a11y-search-input').send_keys(self._vacancy_name)
        self._browser.find_element(
            By.CSS_SELECTOR, '[data-qa="search-button"]').click()

    # Separate links from other objects in selenium element
    def _separate_links_from_selenium_elem(self, selenium_elem_list: list) -> list:
        clean_links_list = []
        for link in selenium_elem_list:
            clean_links_list.append(link.get_attribute('href'))

        return clean_links_list

    def _switch_to_next_vacancy_page(self, page_number: int) -> None:
        self._browser.find_element(
            By.CSS_SELECTOR, f"[data-qa='pager-page-wrapper-{page_number}-{page_number-1}'] [data-qa='pager-page']").click()

    def _list_merge(self, lstlst: list) -> list:
        all = []
        for lst in lstlst:
            all.extend(lst)
        return all

    # Return list with lists with vacancies links
    def _get_vacancy_links(self) -> list:
        vacancy_list = self._browser.find_elements(
            By.CSS_SELECTOR, '[data-qa="serp-item__title"]')
        links_list = []
        links_list.append(
            self._separate_links_from_selenium_elem(vacancy_list))

        # Only 50 vacancies in 1 Head Hunter page. if it is more than 50 vacancies then we should switch to next page with vacancies.
        if self._number_of_vacancy > 50:
            # start from 2 page, because first has already parsed
            for page_number in range(2, self._number_of_vacancy//50+2):
                self._switch_to_next_vacancy_page(page_number)
                vacancy_list = self._browser.find_elements(
                    By.CSS_SELECTOR, '[data-qa="serp-item__title"]')
                links_list.append(
                    self._separate_links_from_selenium_elem(vacancy_list))

        return self._list_merge(links_list)

    # Selenium not so smart. We need manually switch page
    def _switch_to_vacancy_window(self, link: str) -> None:
        self._browser.execute_script(f"window.open('{link}','_blank');")
        new_window = self._browser.window_handles[1]
        self._browser.switch_to.window(new_window)

    # Selenium not so smart. We need manually switch page back
    def _back_to_main_window(self) -> None:
        self._browser.close()
        main_window = self._browser.window_handles[0]
        self._browser.switch_to.window(main_window)

    # Return data list with information from vacancies
    def get_vacancy_information(self) -> None:
        self._fill_start_page()
        vacancy_links = self._get_vacancy_links()[:self._number_of_vacancy]

        self._vacancy_information_list.append(
            ('title', 'salary', 'work_experience', 'vacancy_description', 'company_link', 'skills'))

        # Get information from vacancy page
        for link in vacancy_links:
            self._switch_to_vacancy_window(link)

            title = self._browser.find_element(
                By.CSS_SELECTOR, "[data-qa='vacancy-title']").text
            salary = self._browser.find_element(
                By.CLASS_NAME, 'bloko-header-section-2.bloko-header-section-2_lite').text
            work_experience = self._browser.find_element(
                By.CSS_SELECTOR, "[data-qa='vacancy-experience']").text
            vacancy_description = self._browser.find_element(
                By.CSS_SELECTOR, "[data-qa='vacancy-view-employment-mode']").text
            company_link = self._browser.find_element(
                By.CSS_SELECTOR, "[data-qa='vacancy-company-name']").get_attribute('href')
            skills = self._browser.find_elements(
                By.CSS_SELECTOR, "[data-qa='bloko-tag__text']")

            # merge all skills into one list
            skills = " ,".join([skill.text for skill in skills])

            self._back_to_main_window()
            self._vacancy_information_list.append(
                (title, salary, work_experience, vacancy_description, company_link, skills))

        self._exit()
        return self._vacancy_information_list

    def _exit(self) -> None:
        self._browser.quit()


class Fill_Excel_Table:

    def __init__(self, xl_file_name: str) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.xl_file_name = xl_file_name

    def fill_xl_table(self, vacancy_data: list) -> None:
        # starts from 1, because in xl table countdown from '1'
        for i in range(1, len(vacancy_data)+1):
            self.ws[f'A{i}'] = vacancy_data[i-1][0]
            self.ws[f'B{i}'] = vacancy_data[i-1][1]
            self.ws[f'C{i}'] = vacancy_data[i-1][2]
            self.ws[f'D{i}'] = vacancy_data[i-1][3]
            self.ws[f'E{i}'] = vacancy_data[i-1][4]
            self.ws[f'F{i}'] = vacancy_data[i-1][5]

        self.wb.save(f'{self.xl_file_name}.xlsx')

# Use example
parser1 = HH_Parser('Python', 400)
vacancy_data = parser1.get_vacancy_information()

xl1 = Fill_Excel_Table('Example')
xl1.fill_xl_table(vacancy_data)
